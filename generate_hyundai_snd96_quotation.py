#!/usr/bin/env python3
"""
Hyundai — Saudi National Day 96 City-Color Lighting rental quotation,
in Miradore Experiences branding. Single-sheet, print/PDF-ready, formula-driven.

Pricing: 9 showrooms x 10 city-colors x 15 days @ SAR 125/day = 168,750,
less National Day discount 750 -> net 168,000 excl. VAT; +15% VAT = 193,200.
"""

import math
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import PageSetupProperties

TEAL = "0F6E78"
TEAL_DARK = "0A555D"
TEAL_LIGHT = "E8F2F3"
TEAL_XLIGHT = "F4FAFA"
ORANGE = "F26524"
INK = "1F2A2C"
MUTED = "5B6B6E"
LINE = "C9DADD"
WHITE = "FFFFFF"
YELLOW = "FFF3CD"
FONT = "Arial"
LOGO = "/home/user/miradore/Miradore Logo Color.png"

thin = Side(style="thin", color=LINE)
box = Border(left=thin, right=thin, top=thin, bottom=thin)


def f(size=10, bold=False, color=INK, italic=False):
    return Font(name=FONT, size=size, bold=bold, color=color, italic=italic)


def fill(hex_):
    return PatternFill("solid", fgColor=hex_)


wb = Workbook()
ws = wb.active
ws.title = "Quotation"
NC = 6
widths = [5, 46, 12, 10, 16, 17]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.sheet_view.showGridLines = False


def merge_text(row, text, font_, align="left", height=None, c1=1, c2=NC, fill_hex=None, indent=0):
    ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
    c = ws.cell(row=row, column=c1, value=text)
    c.font = font_
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True, indent=indent)
    if fill_hex:
        for col in range(c1, c2 + 1):
            ws.cell(row=row, column=col).fill = fill(fill_hex)
    if height:
        ws.row_dimensions[row].height = height


def section(row, text):
    ws.cell(row=row, column=1).fill = fill(ORANGE)
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=NC)
    c = ws.cell(row=row, column=2, value=text)
    c.font = f(11.5, True, WHITE)
    c.alignment = Alignment(vertical="center", indent=1)
    for col in range(2, NC + 1):
        ws.cell(row=row, column=col).fill = fill(TEAL)
    ws.row_dimensions[row].height = 22


def th(row, labels, aligns=None):
    for j, lab in enumerate(labels, start=1):
        c = ws.cell(row=row, column=j, value=lab)
        c.font = f(9.5, True, WHITE)
        c.fill = fill(TEAL_DARK)
        c.alignment = Alignment(horizontal=(aligns[j - 1] if aligns else "center"),
                                vertical="center", wrap_text=True)
        c.border = box
    ws.row_dimensions[row].height = 26


# ---------------------------------------------------------------- header
logo = XLImage(LOGO)
scale = 300 / logo.width
logo.width = int(logo.width * scale)
logo.height = int(logo.height * scale)
ws.add_image(logo, "B2")
for r in range(2, 6):
    ws.row_dimensions[r].height = 16

ws.merge_cells("D2:F5")
c = ws["D2"]
c.value = "QUOTATION"
c.font = f(24, True, TEAL)
c.alignment = Alignment(horizontal="right", vertical="center")

ws.row_dimensions[7].height = 3
for col in range(1, NC + 1):
    ws.cell(row=7, column=col).fill = fill(ORANGE)

merge_text(9, "SAUDI NATIONAL DAY 96 — CITY-COLOR LIGHTING", f(15, True, WHITE), height=28, fill_hex=TEAL, indent=1)
merge_text(10, "Detailed BOQ & Commercial Quotation — Hyundai Showrooms, Kingdom of Saudi Arabia", f(9.5, False, WHITE), height=16, fill_hex=TEAL_DARK, indent=1)

# client / ref line
r = 12
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
c = ws.cell(row=r, column=1, value="Client:  Hyundai — Showrooms Network, KSA")
c.font = f(10.5, True)
c.alignment = Alignment(vertical="center", indent=1)
ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=6)
c = ws.cell(row=r, column=4, value="Quotation Ref: MIR-HYN-SND96-001   ·   04 September 2026")
c.font = f(10)
c.alignment = Alignment(horizontal="right", vertical="center", indent=1)
for col in range(1, NC + 1):
    ws.cell(row=r, column=col).fill = fill(TEAL_XLIGHT)
    ws.cell(row=r, column=col).border = box
ws.row_dimensions[r].height = 20

# date chips
r = 13
chips = [
    ((1, 2), "INSTALLATION DATE", "16 September 2026"),
    ((3, 4), "DISMANTLING DATE", "30 September 2026"),
    ((5, 6), "RENTAL DURATION", "15 Days"),
]
ws.row_dimensions[r].height = 15
ws.row_dimensions[r + 1].height = 18
for (c1, c2), lab, val in chips:
    ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
    lc = ws.cell(row=r, column=c1, value=lab)
    lc.font = f(8, True, TEAL)
    lc.alignment = Alignment(horizontal="center", vertical="bottom")
    ws.merge_cells(start_row=r + 1, start_column=c1, end_row=r + 1, end_column=c2)
    vc = ws.cell(row=r + 1, column=c1, value=val)
    vc.font = f(11, True)
    vc.alignment = Alignment(horizontal="center", vertical="top")
    for rr_ in (r, r + 1):
        for col in range(c1, c2 + 1):
            ws.cell(row=rr_, column=col).fill = fill(TEAL_LIGHT)
    # outer border of the chip
    for col in range(c1, c2 + 1):
        ws.cell(row=r, column=col).border = Border(top=thin, left=thin if col == c1 else None,
                                                   right=thin if col == c2 else None)
        ws.cell(row=r + 1, column=col).border = Border(bottom=thin, left=thin if col == c1 else None,
                                                       right=thin if col == c2 else None)

# currency note
r = 15
merge_text(r, "Currency: Saudi Riyal (SAR)   ·   Prices exclude 15% VAT, shown separately in the commercial summary",
           f(8.5, False, MUTED, True), align="center", height=14)
r = 16

# ---------------------------------------------------------------- section 1: location-wise BOQ
r += 1
section(r, "1.  LOCATION-WISE BOQ")
r += 1
th(r, ["#", "Showroom Location", "City-Color\nQty", "Days", "Rate per Unit\n(per day, SAR)", "Total Amount\n(SAR)"])
r += 1
locs = [
    "Jeddah — Auto Mall Showroom",
    "Jeddah — Tahlia Showroom",
    "Jeddah — Al Haramain Showroom",
    "Jeddah — Obhur Showroom",
    "Makkah — Kakia Showroom",
    "Madinah — Airport Road Showroom",
    "Taif — Showroom",
    "Najran — Showroom",
    "Tabuk — Showroom",
]
loc_first = r
for i, loc in enumerate(locs):
    ws.cell(row=r, column=1, value=i + 1).font = f(9.5)
    ws.cell(row=r, column=1).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=r, column=2, value=loc).font = f(9.5)
    ws.cell(row=r, column=2).alignment = Alignment(vertical="center", indent=1)
    q = ws.cell(row=r, column=3, value=10)
    d = ws.cell(row=r, column=4, value=15)
    rt = ws.cell(row=r, column=5, value=160)
    for cell in (q, d, rt):
        cell.font = Font(name=FONT, size=9.5, color="0000FF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    rt.number_format = "#,##0"
    t = ws.cell(row=r, column=6, value=f"=C{r}*D{r}*E{r}")
    t.number_format = "#,##0"
    t.font = f(9.5, True)
    t.alignment = Alignment(horizontal="right", vertical="center")
    for col in range(1, NC + 1):
        ws.cell(row=r, column=col).border = box
        if i % 2 == 1:
            ws.cell(row=r, column=col).fill = fill(TEAL_XLIGHT)
    ws.row_dimensions[r].height = 17
    r += 1
loc_last = r - 1
# total row
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
c = ws.cell(row=r, column=1, value="TOTAL")
c.font = f(10.5, True, WHITE)
c.alignment = Alignment(vertical="center", indent=1)
ws.cell(row=r, column=3, value=f"=SUM(C{loc_first}:C{loc_last})").font = f(10, True, WHITE)
ws.cell(row=r, column=4, value=15).font = f(10, True, WHITE)
ws.cell(row=r, column=5, value="—").font = f(10, True, WHITE)
tt = ws.cell(row=r, column=6, value=f"=SUM(F{loc_first}:F{loc_last})")
tt.number_format = "#,##0"
tt.font = f(10.5, True, WHITE)
tt.alignment = Alignment(horizontal="right", vertical="center")
for col in (3, 4, 5):
    ws.cell(row=r, column=col).alignment = Alignment(horizontal="center", vertical="center")
for col in range(1, NC + 1):
    ws.cell(row=r, column=col).fill = fill(TEAL)
    ws.cell(row=r, column=col).border = box
ws.row_dimensions[r].height = 20
loc_total_row = r

# ---------------------------------------------------------------- section 2: detailed BOQ per showroom
r += 2
section(r, "2.  DETAILED BOQ — PER SHOWROOM")
r += 1
th(r, ["#", "Description", "Qty", "Unit", "Remarks", ""], aligns=["center", "left", "center", "center", "left", "left"])
ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=6)
r += 1
boq = [
    ("City-Color LED Lighting Fixture (Saudi National Day Green)", 10, "Nos", "Rental for 15 days @ SAR 160 per unit per day"),
    ("Transportation to Site", 1, "LS", "Included"),
    ("Loading & Unloading", 1, "LS", "Included"),
    ("Installation of City-Color Fixtures", 10, "Nos", "Included"),
    ("Mounting Brackets & Accessories", 10, "Sets", "Included"),
    ("Electrical Cabling & Connections", 1, "LS", "Included"),
    ("Testing & Commissioning", 1, "LS", "Included"),
    ("Focusing, Aiming & Programming", 1, "LS", "Included"),
    ("Installation Manpower & Supervision", 1, "LS", "Included"),
    ("Technical Support During Rental Period", 1, "LS", "Included"),
    ("Dismantling After Event", 10, "Nos", "Included"),
    ("Removal, Packing & Transportation", 1, "LS", "Included"),
]
for i, (desc, qty, unit, rem) in enumerate(boq):
    ws.cell(row=r, column=1, value=i + 1).font = f(9.5)
    ws.cell(row=r, column=1).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=r, column=2, value=desc).font = f(9.5)
    ws.cell(row=r, column=2).alignment = Alignment(vertical="center", indent=1, wrap_text=True)
    ws.cell(row=r, column=3, value=qty).font = f(9.5)
    ws.cell(row=r, column=3).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=r, column=4, value=unit).font = f(9.5)
    ws.cell(row=r, column=4).alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=6)
    rm = ws.cell(row=r, column=5, value=rem)
    rm.font = f(9, False, MUTED, rem == "Included")
    rm.alignment = Alignment(vertical="center", indent=1, wrap_text=True)
    for col in range(1, NC + 1):
        ws.cell(row=r, column=col).border = box
        if i % 2 == 1:
            ws.cell(row=r, column=col).fill = fill(TEAL_XLIGHT)
    ws.row_dimensions[r].height = 24 if i == 0 else 16
    r += 1

# ---------------------------------------------------------------- section 3: commercial summary
r += 1
section(r, "3.  COMMERCIAL SUMMARY")
r += 1
VATCELL = None
summary = [
    ("Number of Showrooms", f"=COUNTA(B{loc_first}:B{loc_last})", "0", 0),
    ("City-Colors per Showroom", 10, "0", 0),
    ("Total City-Colors", f"=C{loc_total_row}", "0", 0),
    ("Rental Duration", "15 Days", None, 0),
    ("Daily Rate (per City-Color)", f"=E{loc_first}", "#,##0", 0),
    ("Rate per City-Color for 15 Days", f"=E{loc_first}*15", "#,##0", 0),
    ("Cost per Showroom", f"=F{loc_first}", "#,##0", 0),
    ("TOTAL PROJECT VALUE — All 9 Showrooms (excl. VAT)", f"=F{loc_total_row}", "#,##0", 4),
    ("VAT (15%)", None, "#,##0.00", 0),
    ("GRAND TOTAL (incl. VAT)", None, "#,##0.00", 3),
]
sum_first = r
for label, val, fmt, emph in summary:
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    lc = ws.cell(row=r, column=1, value=label)
    lc.alignment = Alignment(vertical="center", indent=1)
    ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=6)
    vc = ws.cell(row=r, column=5)
    if val is not None:
        vc.value = val
    if fmt:
        vc.number_format = fmt
    vc.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    if emph == 3:
        for col in range(1, NC + 1):
            ws.cell(row=r, column=col).fill = fill(ORANGE)
        lc.font = f(11, True, WHITE)
        vc.font = f(12, True, WHITE)
        ws.row_dimensions[r].height = 24
    elif emph == 4:
        for col in range(1, NC + 1):
            ws.cell(row=r, column=col).fill = fill(TEAL_DARK)
        lc.font = f(11.5, True, WHITE)
        vc.font = f(14, True, WHITE)
        ws.row_dimensions[r].height = 28
    elif emph == 1:
        for col in range(1, NC + 1):
            ws.cell(row=r, column=col).fill = fill(TEAL_LIGHT)
        lc.font = f(10, True)
        vc.font = f(10, True)
        ws.row_dimensions[r].height = 18
    elif emph == 2:
        lc.font = f(10)
        vc.font = Font(name=FONT, size=10, bold=True, color="0000FF")
        vc.fill = fill(YELLOW)
        ws.row_dimensions[r].height = 17
    else:
        lc.font = f(10)
        vc.font = f(10)
        ws.row_dimensions[r].height = 17
    for col in range(1, NC + 1):
        ws.cell(row=r, column=col).border = box
    r += 1
# formulas for VAT / grand
total_row = sum_first + 7
vat_row = sum_first + 8
grand_row = sum_first + 9
ws.cell(row=vat_row, column=5, value=f"=ROUND(E{total_row}*0.15,2)")
ws.cell(row=grand_row, column=5, value=f"=E{total_row}+E{vat_row}")

# ---------------------------------------------------------------- section 4: scope
r += 1
section(r, "4.  SCOPE OF SUPPLY & SERVICES (INCLUDED)")
r += 1
merge_text(r, "Equipment Rental   ·   Transportation   ·   Installation   ·   Cabling & Connections   ·   Testing & "
              "Commissioning   ·   Technical Support   ·   Dismantling & Removal",
           f(9.5, True, TEAL), align="center", height=24, fill_hex=TEAL_XLIGHT)
for col in range(1, NC + 1):
    ws.cell(row=r, column=col).border = box
r += 2

# ---------------------------------------------------------------- section 5: notes
section(r, "5.  IMPORTANT NOTES & TERMS")
r += 1
notes = [
    "The rate of SAR 160 per City-Color per day is all-inclusive, covering rental, transportation, installation, accessories, testing, commissioning, technical support, dismantling and removal.",
    "Pricing is based on standard installation conditions. Any special access requirements (crane, boom lift, scaffolding, special permits, major electrical works, etc.) will be quoted separately after site assessment.",
    "Preferred Payment Terms (negotiable): 50% advance upon confirmation / purchase order; 50% upon completion of installation across all showrooms.",
    "Quotation is valid for the rental period stated (16 – 30 September 2026).",
    "All prices are in Saudi Riyals (SAR) and exclude 15% VAT, shown separately above.",
]
for t in notes:
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NC)
    c = ws.cell(row=r, column=1, value="•  " + t)
    c.font = f(9)
    c.alignment = Alignment(wrap_text=True, vertical="top", indent=1)
    lines = max(1, math.ceil(len(t) / 100))
    ws.row_dimensions[r].height = lines * 12 + 5
    r += 1

# ---------------------------------------------------------------- signature & stamp
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils.units import pixels_to_EMU

r += 1
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
c = ws.cell(row=r, column=1, value="For & on behalf of — MIRADORE EXPERIENCES")
c.font = f(10, True, TEAL)
c.alignment = Alignment(vertical="center", indent=1)
ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=6)
c = ws.cell(row=r, column=4, value="CLIENT ACCEPTANCE — HYUNDAI")
c.font = f(10, True, TEAL)
c.alignment = Alignment(vertical="center", indent=1)
ws.row_dimensions[r].height = 18
sig_top = r + 1
for rr in range(sig_top, sig_top + 6):
    ws.row_dimensions[rr].height = 17

def place(path, px_w, px_h, col, col_off_px, row, row_off_px):
    im = XLImage(path)
    marker = AnchorMarker(col=col, colOff=pixels_to_EMU(col_off_px),
                          row=row, rowOff=pixels_to_EMU(row_off_px))
    im.anchor = OneCellAnchor(_from=marker,
                              ext=XDRPositiveSize2D(pixels_to_EMU(px_w), pixels_to_EMU(px_h)))
    ws.add_image(im)

place("/home/user/miradore/miradore_signature.png", 215, 60, col=1, col_off_px=8, row=sig_top - 1, row_off_px=26)
place("/home/user/miradore/miradore_stamp.png", 133, 103, col=1, col_off_px=140, row=sig_top - 1, row_off_px=2)

r = sig_top + 6
top_line = Border(top=Side(style="thin", color=MUTED))
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
c = ws.cell(row=r, column=1, value="Adeel Ahmed — Director")
c.font = f(9.5, True)
c.alignment = Alignment(vertical="center", indent=1)
ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=6)
c = ws.cell(row=r, column=4, value="Signature, Company Stamp & Date")
c.font = f(9, False, MUTED)
c.alignment = Alignment(vertical="center", indent=1)
for col in (1, 2, 4, 5, 6):
    ws.cell(row=r, column=col).border = top_line
ws.row_dimensions[r].height = 16
r += 1
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
c = ws.cell(row=r, column=1, value="Authorized Signatory")
c.font = f(8.5, False, MUTED)
c.alignment = Alignment(vertical="top", indent=1)
ws.row_dimensions[r].height = 14

r += 2
for col in range(1, NC + 1):
    ws.cell(row=r, column=col).fill = fill(TEAL)
    ws.cell(row=r + 1, column=col).fill = fill(TEAL)
ws.merge_cells(start_row=r, start_column=1, end_row=r + 1, end_column=NC)
c = ws.cell(row=r, column=1, value="MIRADORE EXPERIENCES — RIYADH, KINGDOM OF SAUDI ARABIA\nLighting a Brighter Tomorrow for Saudi Arabia  ·  Together We Shine")
c.font = f(9, True, WHITE)
c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
end = r + 1

# ---------------------------------------------------------------- page setup
ws.page_setup.orientation = "portrait"
ws.page_setup.paperSize = 9
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 1
ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
ws.page_margins.left = ws.page_margins.right = 0.45
ws.page_margins.top = 0.5
ws.page_margins.bottom = 0.55
ws.oddFooter.center.text = "Miradore Experiences — Riyadh, KSA   |   Page &P of &N"
ws.oddFooter.center.size = 8
ws.oddFooter.center.font = FONT
ws.oddFooter.center.color = MUTED
ws.print_area = f"A1:F{end}"

wb.properties.title = "Hyundai SND96 City-Color Lighting — Quotation — Miradore Experiences"
wb.properties.creator = "Miradore Experiences"
out = "/home/user/miradore/Hyundai_SND96_CityColor_Quotation.xlsx"
wb.save(out)
print("saved", out)
