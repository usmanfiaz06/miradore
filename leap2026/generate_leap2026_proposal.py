#!/usr/bin/env python3
"""
PSEB — Pakistan Saudi Business Forum & Pakistan Pavilion at LEAP 2026
Commercial Proposal workbook generator for Miradore Experiences, Riyadh.

Builds one print-ready Excel workbook (exportable straight to PDF):
  Cover | Summary | Lot A - Forum | Lot B - Pavilion | Team | Internal (hidden)

All money cells are formulas driven by editable inputs (rates, qty, VAT %,
agency commission %, SAR->PKR rate) so the numbers stay live.

Pricing structure:
  - Crowne Plaza venue & catering package (Lot A / A.1): 325,000 SAR + VAT,
    passed through at actual cost (no markup, no commission).
  - Lot A production & services: supplier BOQ prices marked up per line.
  - Lot B turnkey pavilion: single all-inclusive package at 232,000 + VAT.
  - Agency commission 12% on production & service items, per lot.
"""

import math
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import PageSetupProperties

# ----------------------------------------------------------------------------- brand
TEAL = "0F6E78"
TEAL_DARK = "0A555D"
TEAL_LIGHT = "E8F2F3"
TEAL_XLIGHT = "F4FAFA"
ORANGE = "F26524"
INK = "1F2A2C"
MUTED = "5B6B6E"
LINE = "C9DADD"
WHITE = "FFFFFF"
YELLOW = "FFF3CD"

FONT = "Arial"
LOGO = "/home/user/miradore/Miradore Logo Color.png"

thin = Side(style="thin", color=LINE)
box = Border(left=thin, right=thin, top=thin, bottom=thin)


def f(size=10, bold=False, color=INK, italic=False):
    return Font(name=FONT, size=size, bold=bold, color=color, italic=italic)


def fill(hex_):
    return PatternFill("solid", fgColor=hex_)


def set_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def band(ws, row, ncols, text, sub=None):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = f(15, True, WHITE)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    for col in range(1, ncols + 1):
        ws.cell(row=row, column=col).fill = fill(TEAL)
    ws.row_dimensions[row].height = 30
    if sub:
        ws.merge_cells(start_row=row + 1, start_column=1, end_row=row + 1, end_column=ncols)
        s = ws.cell(row=row + 1, column=1, value=sub)
        s.font = f(9.5, False, WHITE)
        s.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        for col in range(1, ncols + 1):
            ws.cell(row=row + 1, column=col).fill = fill(TEAL_DARK)
        ws.row_dimensions[row + 1].height = 17


def desc_height(text, cpl=62, line_h=12.5, extra=6, min_h=20):
    lines = 0
    for seg in str(text).split("\n"):
        lines += max(1, math.ceil(len(seg) / cpl))
    return max(min_h, lines * line_h + extra)


def page(ws, landscape=False, fit_h=1, titles=None, area=None):
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape" if landscape else "portrait"
    ws.page_setup.paperSize = 9  # A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = fit_h
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_margins.left = ws.page_margins.right = 0.45
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.55
    ws.oddFooter.center.text = "Miradore Experiences — Riyadh, KSA   |   Page &P of &N"
    ws.oddFooter.center.size = 8
    ws.oddFooter.center.font = FONT
    ws.oddFooter.center.color = MUTED
    if titles:
        ws.print_title_rows = titles
    if area:
        ws.print_area = area


wb = Workbook()

# =============================================================================
# 1. COVER
# =============================================================================
cv = wb.active
cv.title = "Cover"
set_widths(cv, [4, 15, 15, 15, 15, 15, 15, 4])
NC = 8

for r in range(1, 3):
    for col in range(1, NC + 1):
        cv.cell(row=r, column=col).fill = fill(TEAL)
cv.row_dimensions[1].height = 8
cv.row_dimensions[2].height = 8

logo = XLImage(LOGO)
scale = 430 / logo.width
logo.width = int(logo.width * scale)
logo.height = int(logo.height * scale)
cv.add_image(logo, "C5")
for r in range(4, 11):
    cv.row_dimensions[r].height = 18

def cover_line(row, text, font_, align="center", height=None, col_start=2, col_end=7):
    cv.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
    c = cv.cell(row=row, column=col_start, value=text)
    c.font = font_
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    if height:
        cv.row_dimensions[row].height = height

cover_line(12, "COMMERCIAL PROPOSAL", f(26, True, TEAL), height=36)
cover_line(13, "(Financial Proposal — Lot A & Lot B)", f(11, False, MUTED), height=16)
cover_line(15, "Pakistan Saudi Business Forum", f(18, True, INK), height=26)
cover_line(16, "& Pakistan Pavilion at LEAP 2026", f(18, True, INK), height=26)
cover_line(17, "Riyadh, Kingdom of Saudi Arabia   ·   30 August – 3 September 2026", f(11.5, False, MUTED), height=20)

cv.row_dimensions[19].height = 3
for col in range(3, 7):
    cv.cell(row=19, column=col).fill = fill(ORANGE)

cover_line(21, "PREPARED FOR", f(10, True, ORANGE), height=16)
cover_line(22, "Pakistan Software Export Board (PSEB)", f(13, True, INK), height=20)
cover_line(23, "Ministry of IT & Telecommunication, Government of Pakistan", f(10.5, False, MUTED), height=15)
cover_line(24, "6th Floor, New State Life Tower, Jinnah Avenue, Blue Area, Islamabad", f(10.5, False, MUTED), height=15)

cover_line(26, "In response to PSEB RFQ — Onboarding an Event Management Vendor for the Pakistan Saudi Business\nForum and the Pakistan Pavilion at LEAP 2026 in Riyadh, Kingdom of Saudi Arabia", f(9.5, False, MUTED, True), height=28)

cover_line(29, "PREPARED BY", f(10, True, ORANGE), height=16)
cover_line(30, "Miradore Experiences — Riyadh, Kingdom of Saudi Arabia", f(13, True, INK), height=20)

cover_line(32, "Ref: MIR-PSEB-LEAP-2026-001        Date: 21 July 2026        Validity: 60 days", f(10, True, INK), height=16)

for r in range(35, 37):
    for col in range(1, NC + 1):
        cv.cell(row=r, column=col).fill = fill(TEAL)
cv.merge_cells(start_row=35, start_column=1, end_row=36, end_column=NC)
c = cv.cell(row=35, column=1, value="End-to-End Event Management  ·  Turnkey Pavilion Construction  ·  Premium Branding & AV  ·  Media Production")
c.font = f(9.5, True, WHITE)
c.alignment = Alignment(horizontal="center", vertical="center")

page(cv, area="A1:H36")

# =============================================================================
# 2. SUMMARY  (hosts the 3 commercial parameters used everywhere)
# =============================================================================
sm = wb.create_sheet("Summary")
set_widths(sm, [3, 46, 17, 17, 20, 3])
NC = 6
band(sm, 1, NC, "EXECUTIVE SUMMARY & COMMERCIAL OVERVIEW",
     "Pakistan Saudi Business Forum (Lot A) · Pakistan Pavilion at LEAP 2026 (Lot B) · Riyadh, KSA")

sm.merge_cells("B4:E4")
c = sm["B4"]
c.value = ("Miradore Experiences is pleased to submit this financial proposal for both lots of the PSEB RFQ. "
           "Lot A covers the 300-guest Pakistan Saudi Business Forum at Crowne Plaza Riyadh RDC — the hotel venue & "
           "catering package is passed through at actual cost, with all production, branding, AV, staging and media "
           "services itemized at unit rates. Lot B covers the 162-SQM Pakistan Pavilion at Riyadh Exhibition & "
           "Convention Centre, Malham (Hall 3 — H170 / J170), offered as a single all-inclusive turnkey package "
           "(design, fabrication, approvals, power and installation) plus daily hospitality, provisioning, cleaning "
           "and multi-crew media coverage across all four LEAP days.")
c.font = f(10)
c.alignment = Alignment(wrap_text=True, vertical="top")
sm.row_dimensions[4].height = 78

# ---- Commercial parameters (single source of truth)
PR = 6
sm.merge_cells(start_row=PR, start_column=2, end_row=PR, end_column=5)
c = sm.cell(row=PR, column=2, value="COMMERCIAL PARAMETERS  (editable inputs)")
c.font = f(10, True, WHITE)
c.alignment = Alignment(vertical="center", indent=1)
for col in range(2, 6):
    sm.cell(row=PR, column=col).fill = fill(TEAL)
sm.row_dimensions[PR].height = 18

params = [
    ("VAT (Kingdom of Saudi Arabia)", 0.15, "0%", "Applied on all services"),
    ("Agency Commission / Management Fee", 0.12, "0%", "Applied on each lot subtotal"),
    ("Exchange rate — PKR per 1 SAR (reference)", 75.50, "#,##0.00", "Update on submission date; drives all PKR equivalents"),
]
for i, (label, val, fmt, note) in enumerate(params):
    r = PR + 1 + i
    sm.cell(row=r, column=2, value=label).font = f(10)
    v = sm.cell(row=r, column=3, value=val)
    v.font = Font(name=FONT, size=10, bold=True, color="0000FF")
    v.fill = fill(YELLOW)
    v.number_format = fmt
    v.alignment = Alignment(horizontal="center")
    sm.merge_cells(start_row=r, start_column=4, end_row=r, end_column=5)
    n = sm.cell(row=r, column=4, value=note)
    n.font = f(9, False, MUTED, True)
    n.alignment = Alignment(vertical="center")
    for col in range(2, 6):
        sm.cell(row=r, column=col).border = box
    sm.row_dimensions[r].height = 17

VAT = "Summary!$C$7"
COMM = "Summary!$C$8"
FX = "Summary!$C$9"

# Lot sheet footer cells (built below; asserted to match)
LOTA, LOTB = "'Lot A - Forum'", "'Lot B - Pavilion'"
LA = {"sub": "F20", "comm": "F21", "ex": "F22", "vat": "F23", "inc": "F24"}
LB = {"sub": "F11", "comm": "F12", "ex": "F13", "vat": "F14", "inc": "F15"}

TR = PR + 5  # 11
sm.merge_cells(start_row=TR, start_column=2, end_row=TR, end_column=5)
c = sm.cell(row=TR, column=2, value="COMMERCIAL SUMMARY")
c.font = f(11, True, TEAL)
sm.row_dimensions[TR].height = 20

hdr = ["Description", "Amount (SAR)", "Amount (PKR)", "Notes"]
r = TR + 1
for j, htxt in enumerate(hdr):
    c = sm.cell(row=r, column=2 + j, value=htxt)
    c.font = f(9.5, True, WHITE)
    c.fill = fill(TEAL)
    c.alignment = Alignment(horizontal="center" if j else "left", vertical="center", indent=0 if j else 1)
    c.border = box
sm.row_dimensions[r].height = 18

rows = [
    ("Lot A — Pakistan Saudi Business Forum (30 Aug 2026)", f"={LOTA}!{LA['sub']}", "Incl. Crowne Plaza venue & catering package at cost (325,000) + itemized production & services", 28),
    ("Lot B — Pakistan Pavilion at LEAP 2026 (31 Aug – 3 Sep)", f"={LOTB}!{LB['sub']}", "Turnkey pavilion package + LEAP-days services — see Lot B sheet", 28),
    ("Combined Subtotal", None, "", 18),
    ("Agency Commission / Management Fee", None, "As per Commercial Parameters", 18),
    ("Total (excl. VAT)", None, "", 18),
    ("VAT", None, "", 18),
    ("GRAND TOTAL — BOTH LOTS (incl. VAT)", None, "Inclusive of all applicable taxes", 22),
]
R0 = r + 1
for i, (label, formula, note, ht) in enumerate(rows):
    rr = R0 + i
    lab = sm.cell(row=rr, column=2, value=label)
    sar = sm.cell(row=rr, column=3)
    pkr = sm.cell(row=rr, column=4)
    nt = sm.cell(row=rr, column=5, value=note)
    if formula:
        sar.value = formula
    lab.font = f(10)
    lab.alignment = Alignment(vertical="center", wrap_text=True)
    nt.font = f(8.5, False, MUTED, True)
    nt.alignment = Alignment(vertical="center", wrap_text=True)
    sar.number_format = "#,##0"
    pkr.number_format = "#,##0"
    sar.alignment = pkr.alignment = Alignment(horizontal="right", vertical="center")
    for col in range(2, 6):
        sm.cell(row=rr, column=col).border = box
    sm.row_dimensions[rr].height = ht

sm.cell(row=R0 + 2, column=3, value=f"=C{R0}+C{R0+1}")
sm.cell(row=R0 + 3, column=3, value=f"={LOTA}!{LA['comm']}+{LOTB}!{LB['comm']}")
sm.cell(row=R0 + 4, column=3, value=f"=C{R0+2}+C{R0+3}")
sm.cell(row=R0 + 5, column=3, value=f"=ROUND(C{R0+4}*{VAT},2)")
sm.cell(row=R0 + 6, column=3, value=f"=C{R0+4}+C{R0+5}")
for i in range(7):
    sm.cell(row=R0 + i, column=4, value=f"=ROUND(C{R0+i}*{FX},0)")
for i in (2, 4):
    for col in range(2, 6):
        sm.cell(row=R0 + i, column=col).fill = fill(TEAL_LIGHT)
    for col in (2, 3, 4):
        sm.cell(row=R0 + i, column=col).font = f(10, True)
gr = R0 + 6
for col in range(2, 6):
    sm.cell(row=gr, column=col).fill = fill(TEAL)
sm.cell(row=gr, column=2).font = f(11, True, WHITE)
for col in (3, 4):
    sm.cell(row=gr, column=col).font = f(11, True, WHITE)
    sm.cell(row=gr, column=col).number_format = "#,##0.00" if col == 3 else "#,##0"
sm.cell(row=gr, column=5).font = f(8.5, False, WHITE, True)

# ---- Per-lot standalone totals
TR2 = gr + 2
sm.merge_cells(start_row=TR2, start_column=2, end_row=TR2, end_column=5)
sm.cell(row=TR2, column=2, value="LOT-WISE TOTALS  (each lot priced to stand alone for independent award)").font = f(11, True, TEAL)
hdr2 = ["Lot", "Total excl. VAT (SAR)", "Total incl. VAT (SAR)", "Total incl. VAT (PKR)"]
r = TR2 + 1
for j, htxt in enumerate(hdr2):
    c = sm.cell(row=r, column=2 + j, value=htxt)
    c.font = f(9.5, True, WHITE)
    c.fill = fill(TEAL)
    c.alignment = Alignment(horizontal="center" if j else "left", vertical="center", indent=0 if j else 1)
    c.border = box
sm.row_dimensions[r].height = 18
lots = [
    ("Lot A — Business Forum (incl. agency commission)", f"={LOTA}!{LA['ex']}", f"={LOTA}!{LA['inc']}"),
    ("Lot B — LEAP Pavilion (incl. agency commission)", f"={LOTB}!{LB['ex']}", f"={LOTB}!{LB['inc']}"),
]
for i, (label, ex, inc) in enumerate(lots):
    rr = r + 1 + i
    sm.cell(row=rr, column=2, value=label).font = f(10)
    sm.cell(row=rr, column=2).alignment = Alignment(vertical="center", wrap_text=True)
    sm.cell(row=rr, column=3, value=ex)
    sm.cell(row=rr, column=4, value=inc)
    sm.cell(row=rr, column=5, value=f"=ROUND(D{rr}*{FX},0)")
    for col in (3, 4):
        sm.cell(row=rr, column=col).number_format = "#,##0.00"
        sm.cell(row=rr, column=col).alignment = Alignment(horizontal="right", vertical="center")
    sm.cell(row=rr, column=5).number_format = "#,##0"
    sm.cell(row=rr, column=5).alignment = Alignment(horizontal="right", vertical="center")
    for col in range(2, 6):
        sm.cell(row=rr, column=col).border = box
    sm.row_dimensions[rr].height = 20

# ---- Terms
TT = r + 4
sm.merge_cells(start_row=TT, start_column=2, end_row=TT, end_column=5)
sm.cell(row=TT, column=2, value="PAYMENT TERMS & KEY CONDITIONS").font = f(11, True, TEAL)
terms = [
    "Preferred Payment Terms (negotiable): 50% advance upon confirmation / work order — primarily to cover the mandatory advance payment required by Crowne Plaza Riyadh RDC to secure the venue, plus mobilization of fabrication and long-lead items; 30% two (2) days before the event; 20% balance after the event, upon successful delivery and verification.",
    "The Crowne Plaza Riyadh RDC venue & catering package (Lot A, item A.1) is passed through at actual hotel package cost, with no markup.",
    "Prices are quoted in Saudi Riyals (SAR); PKR equivalents are provided at the reference exchange rate stated above and will be aligned to the prevailing rate at invoicing.",
    "Quoted totals are inclusive of all applicable taxes; VAT (15%) is shown separately for full transparency.",
    "PSEB may increase or decrease quantities of any item; unit rates above will apply pro-rata.",
    "All content, creative assets and campaign data produced under either lot remain the exclusive intellectual property of PSEB; no material will be disclosed or published without PSEB's prior written approval.",
    "A dedicated Project Manager is assigned exclusively to PSEB as single point of contact (see Team & Resource Plan).",
    "Menu options for the 5-course continental buffet will be provided for PSEB approval upon award.",
    "Validity: 60 days from the date of submission.",
]
rr = TT + 1
for t in terms:
    sm.merge_cells(start_row=rr, start_column=2, end_row=rr, end_column=5)
    c = sm.cell(row=rr, column=2, value="•  " + t)
    c.font = f(9.5)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    sm.row_dimensions[rr].height = desc_height(t, cpl=105, line_h=12, extra=4, min_h=15)
    rr += 1

page(sm, area=f"A1:F{rr+1}")

# =============================================================================
# BOQ helpers
# =============================================================================
BOQ_NC = 8

def boq_header(ws, title, subtitle):
    set_widths(ws, [10, 62, 7, 9, 13, 14, 14, 16])
    band(ws, 1, BOQ_NC, title, subtitle)
    HR = 4
    headers = ["Item\nCode", "Description of Sourced Scope", "Qty", "Unit",
               "Unit Rate\n(SAR)", "Total\n(SAR)", "Unit Rate\n(PKR)", "Total\n(PKR)"]
    for j, htxt in enumerate(headers, start=1):
        c = ws.cell(row=HR, column=j, value=htxt)
        c.font = f(9.5, True, WHITE)
        c.fill = fill(TEAL)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = box
    ws.row_dimensions[HR].height = 30
    return HR


def item_row(ws, r, code, desc, qty, unit, rate, fmt, shade):
    ws.cell(row=r, column=1, value=code).font = f(10, True, TEAL)
    ws.cell(row=r, column=1).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    d = ws.cell(row=r, column=2, value=desc)
    d.font = f(9.5)
    d.alignment = Alignment(wrap_text=True, vertical="center")
    if rate is None:
        # included-in-package row: no amount, excluded from sums (text is ignored by SUM)
        ws.cell(row=r, column=3, value="—")
        ws.cell(row=r, column=4, value="—")
        inc = ws.cell(row=r, column=6, value="Included in A.1")
        inc.font = f(8.5, False, MUTED, True)
        ws.cell(row=r, column=7, value="—").font = f(9.5, False, MUTED)
        ws.cell(row=r, column=8, value="—").font = f(9.5, False, MUTED)
        for col in (3, 4, 6, 7, 8):
            ws.cell(row=r, column=col).alignment = Alignment(horizontal="center", vertical="center")
        for col in (3, 4):
            ws.cell(row=r, column=col).font = f(9.5)
    else:
        ws.cell(row=r, column=3, value=qty).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=r, column=3).font = f(9.5)
        ws.cell(row=r, column=4, value=unit).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=r, column=4).font = f(9.5)
        rc = ws.cell(row=r, column=5, value=rate)
        rc.number_format = fmt
        rc.font = Font(name=FONT, size=9.5, color="0000FF")
        rc.alignment = Alignment(horizontal="right", vertical="center")
        tc = ws.cell(row=r, column=6, value=f"=ROUND(C{r}*E{r},2)")
        tc.number_format = "#,##0"
        tc.font = f(9.5, True)
        tc.alignment = Alignment(horizontal="right", vertical="center")
        pr_ = ws.cell(row=r, column=7, value=f"=ROUND(E{r}*{FX},0)")
        pt = ws.cell(row=r, column=8, value=f"=ROUND(F{r}*{FX},0)")
        for cc in (pr_, pt):
            cc.number_format = "#,##0"
            cc.font = f(9.5, False, MUTED)
            cc.alignment = Alignment(horizontal="right", vertical="center")
    if shade:
        for col in range(1, BOQ_NC + 1):
            ws.cell(row=r, column=col).fill = fill(TEAL_XLIGHT)
    for col in range(1, BOQ_NC + 1):
        ws.cell(row=r, column=col).border = box
    ws.row_dimensions[r].height = desc_height(desc)


def foot(ws, row, label, sar_formula, emph=0):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    c = ws.cell(row=row, column=1, value=label)
    c.alignment = Alignment(horizontal="right", vertical="center")
    sc = ws.cell(row=row, column=6, value=sar_formula)
    sc.number_format = "#,##0.00"
    sc.alignment = Alignment(horizontal="right", vertical="center")
    pk = ws.cell(row=row, column=8, value=f"=ROUND(F{row}*{FX},0)")
    pk.number_format = "#,##0"
    pk.alignment = Alignment(horizontal="right", vertical="center")
    for col in range(1, BOQ_NC + 1):
        ws.cell(row=row, column=col).border = box
    if emph == 2:
        for col in range(1, BOQ_NC + 1):
            ws.cell(row=row, column=col).fill = fill(TEAL)
        c.font = f(11, True, WHITE)
        sc.font = f(11, True, WHITE)
        pk.font = f(10, True, WHITE)
        ws.row_dimensions[row].height = 22
    elif emph == 1:
        for col in range(1, BOQ_NC + 1):
            ws.cell(row=row, column=col).fill = fill(TEAL_LIGHT)
        c.font = f(10, True)
        sc.font = f(10, True)
        pk.font = f(9.5, True, MUTED)
        ws.row_dimensions[row].height = 18
    else:
        c.font = f(10)
        sc.font = f(10)
        pk.font = f(9.5, False, MUTED)
        ws.row_dimensions[row].height = 17


def notes_block(ws, start_row, note_lines):
    rr = start_row
    for t in note_lines:
        ws.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=BOQ_NC)
        c = ws.cell(row=rr, column=1, value=t)
        c.font = f(8.5, False, MUTED, True)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[rr].height = desc_height(t, cpl=150, line_h=11, extra=3, min_h=13)
        rr += 1
    return rr


# =============================================================================
# 3. LOT A — items: (code, desc, qty, unit, rate, fmt); rate None => included in A.1
# =============================================================================
la_ws = wb.create_sheet("Lot A - Forum")
HR = boq_header(la_ws, "LOT A — PAKISTAN SAUDI BUSINESS FORUM",
                "Crowne Plaza Riyadh RDC · 30 August 2026 · 300 Guests · Itemized BOQ (SAR, with PKR equivalents)")

lota_items = [
    ("A.1", "Venue Booking for 30 Aug 2026 — Crowne Plaza Riyadh RDC Hotel: booking on PSEB's behalf and securing the "
            "venue by facilitating all mandatory advance payments required by hotel management. All-inclusive hotel "
            "package covering ballroom/hall space for 300 guests, 5-course continental buffet dinner with two "
            "coffee-break services, banquet seating & tables, hotel service staff and dedicated 5 Mbps internet. "
            "Passed through at actual hotel package cost.", 1, "Package", 325000, "#,##0"),
    ("A.1.1", "Event Management, Production Planning, Venue Coordination & Technical Rehearsal — end-to-end show "
              "management for the Forum.", 1, "Job", 18500, "#,##0"),
    ("A.1.2", "Main Stage — portable stage 10m x 4m (40 sqm) with premium stage carpet and step access, as part of "
              "the main-stage arrangements under A.1.", 1, "Job", 12500, "#,##0"),
    ("A.2", "VVIP Front-Row Premium Sofa Seating Layout with personalised name tags — 50 premium sofas, including 25 "
            "coffee tables with black stretch cloth and floral table-top arrangements.", 1, "Lot", 19500, "#,##0"),
    ("A.3", "10ft x 8ft Media Wall (flex with stand) outside the conference hall, with red carpet (12m x 2m), ambiance "
            "lighting and 6 Q-poles for entrance styling.", 1, "Job", 10500, "#,##0"),
    ("A.4", "Reception / Guest Registration Desk with 2 chairs and flex backdrop branding.", 1, "Job", 13500, "#,##0"),
    ("A.5", "High-Luminance 55\" Digital Standees around the greeting zone, including electricity sourcing and "
            "PSEB-approved digital content loop (content prepared by vendor).", 4, "Nos", 1350, "#,##0"),
    ("A.6", "Complete Lighting & Sound System for keynotes, presentations and panel discussions — line-array speakers, "
            "digital console, wireless and podium microphones, moving heads, parcans, city colors and uplights, with "
            "dedicated audio and lighting engineers.", 1, "Job", 26500, "#,##0"),
    ("A.7", "High-Resolution 30ft x 12ft SMD Main Screen with integrated neon border, including media server, washout "
            "and full content design as per PSEB-approved event flow and speakers.", 1, "Job", 34500, "#,##0"),
    ("A.7.1", "6ft x 8ft SMD Mirror Panels throughout the hall, with content designed as per PSEB's directions and "
              "approval.", 4, "Nos", 3150, "#,##0"),
    ("A.8", "Customised wooden speaker podium with the official event logo integrated onto the front fascia, podium "
            "microphone and step stool.", 1, "Job", 5000, "#,##0"),
    ("A.9", "5-Course Continental Buffet Dinner for 300 participants (buffet setup) with table service for the "
            "Minister and VVIP tables — included within the Crowne Plaza hotel package under A.1; menu options to be "
            "provided for PSEB approval.", None, None, None, None),
    ("A.10", "Dedicated on-site project coordinator liaising continuously with hotel banquet staff to ensure flawless "
             "event flow throughout the Forum.", 1, "Job", 5500, "#,##0"),
    ("A.11", "Sourced Media Crew — one professional photography team and one videography team for complete event "
             "coverage, including: 100% raw photo/video upload to a shared secure Google Drive by midnight; polished "
             "90-second highlight reel next day; ten short speaker/segment clips (15-sec and 30-sec variants) for "
             "social media next day; and a 3–5 minute high-production event documentary within 2 days of the Forum.",
     1, "Job", 11500, "#,##0"),
    ("A.12", "Professional Ushers / Hostesses for guest management, registration support and VVIP handling.",
     10, "Person", 950, "#,##0"),
]
r = HR + 1
first = r
for i, (code, desc, qty, unit, rate, fmt) in enumerate(lota_items):
    item_row(la_ws, r, code, desc, qty, unit, rate, fmt, shade=(i % 2 == 1))
    r += 1
last = r - 1

foot(la_ws, r,     "Subtotal", f"=SUM(F{first}:F{last})", emph=1)
foot(la_ws, r + 1, "Agency Commission / Management Fee (12%)", f"=ROUND(F{r}*{COMM},2)")
foot(la_ws, r + 2, "Total (excl. VAT)", f"=F{r}+F{r+1}", emph=1)
foot(la_ws, r + 3, "VAT (15%)", f"=ROUND(F{r+2}*{VAT},2)")
foot(la_ws, r + 4, "LOT A — TOTAL (incl. VAT)", f"=F{r+2}+F{r+3}", emph=2)
la_refs = {"sub": f"F{r}", "comm": f"F{r+1}", "ex": f"F{r+2}", "vat": f"F{r+3}", "inc": f"F{r+4}"}

lota_notes = [
    "Note 1: A.1 is the Crowne Plaza Riyadh RDC all-inclusive package (hall space + full catering for 300 guests) passed through at actual hotel package cost, with no markup.",
    "Note 2: A.9 (5-course continental buffet & coffee breaks) is included within the A.1 hotel package; final headcount will be settled on actuals as per PSEB confirmation.",
    "Note 3: All artwork and screen content will be produced by Miradore and released only after PSEB's written approval.",
]
end = notes_block(la_ws, r + 7, lota_notes)
page(la_ws, landscape=True, fit_h=1, titles=f"{HR}:{HR}", area=f"A1:H{end+1}")

# =============================================================================
# 4. LOT B — turnkey pavilion as ONE package line + LEAP-days services
# =============================================================================
lb_ws = wb.create_sheet("Lot B - Pavilion")
HR = boq_header(lb_ws, "LOT B — PAKISTAN PAVILION AT LEAP 2026",
                "Riyadh Exhibition & Convention Centre (RECC), Malham · 31 Aug – 3 Sep 2026 · 162 SQM · Hall 3 (H170 / J170)")

turnkey_desc = (
    "TURNKEY PAKISTAN PAVILION — ALL-INCLUSIVE PACKAGE (162 SQM · 3 Blocks · 20 Exhibitor Booths · 30-SQM Networking "
    "Lounge · Ministerial Meeting Room), built to Annex I–III blueprints & specifications. Package includes:\n"
    "•  Design, fabrication, sourcing, logistics & transportation to RECC Malham; structural build of Blocks A, B & C; "
    "on-site management and full dismantling\n"
    "•  Block A & C wood/MDF build (16 booths: 2 x 9-SQM + 6 x 6-SQM per block); Block B premium architectural build "
    "(4 x 6-SQM booths, lounge & meeting room)\n"
    "•  Glass-walled ministerial meeting room (two 250cm x 240cm glass walls + glass door) with 55\" LED TV, 2 "
    "single-seater and one 2-seater sofa\n"
    "•  P2.6 LED screens (3.5m x 1m) for block branding; 43\" LED TV per booth; backlit light boxes; neon accent "
    "lighting; premium wooden flooring with neon edge\n"
    "•  3D acrylic + neon block logos, die-cut vinyl logos and full booth branding\n"
    "•  Branded info counter with bar stool, 3 chairs and 1 round table per booth; networking lounge seating with "
    "round tables\n"
    "•  Organizer stand-build permit, official stand audit approval, structural clearance and health/safety/security "
    "clearances ahead of move-in deadlines\n"
    "•  Main electrical power line drops, panel deployment and per-booth hookups — minimum 2 multi-pin power sockets "
    "in every booth, meeting room and networking area, engineered to run all SMDs, LCDs and lighting rigs"
)

lotb_items = [
    ("B.1 – B.6", turnkey_desc, 1, "Package", 232000, "#,##0"),
    ("B.7", "Running Tea & Coffee Station for the entire pavilion (approx. 500 cups/day) with service staff and "
            "consumables — 4 days.", 4, "Day", 4250, "#,##0"),
    ("B.8", "Lounge Catering — premium assorted dates, traditional one-bite snacks and local confectioneries served in "
            "traditional plates daily (approx. 200 person-servings/day), with mineral water for the networking lounge "
            "— 4 days.", 4, "Day", 4800, "#,##0"),
    ("B.9", "Daily Stall Provisions — 6 small chilled mineral water bottles delivered to every booth at the start of "
            "each morning session, plus 100 bottles/day for the lounge (approx. 880 bottles over 4 days).",
     880, "Bottle", 3.5, "#,##0.00"),
    ("B.10", "Continuous Sanitation, Waste Removal & Booth Dustbin Deployment — individual dustbin in each of the 20 "
             "booths emptied routinely, with dedicated daily cleaning of all booths and lounge area — 4 days.",
     4, "Day", 1850, "#,##0"),
    ("B.11", "Sourced Media Crew — 2 dedicated photography teams and 2 videography teams, permanent full-day presence "
             "for 4 days covering ministerial bilaterals, speaker sessions, booth engagement and visitor traffic. "
             "Daily deliverables: 100% raw uploads to shared Drive by midnight; 90-second daily highlight video by "
             "next morning; minimum 10 short clips daily (15-sec / 30-sec) by 10 AM; and 3–5 minute pavilion "
             "documentary within 2 days of event close.", 4, "Day", 9750, "#,##0"),
]
r = HR + 1
first = r
for i, (code, desc, qty, unit, rate, fmt) in enumerate(lotb_items):
    item_row(lb_ws, r, code, desc, qty, unit, rate, fmt, shade=(i % 2 == 1))
    r += 1
last = r - 1

foot(lb_ws, r,     "Subtotal", f"=SUM(F{first}:F{last})", emph=1)
foot(lb_ws, r + 1, "Agency Commission / Management Fee (12%)", f"=ROUND(F{r}*{COMM},2)")
foot(lb_ws, r + 2, "Total (excl. VAT)", f"=F{r}+F{r+1}", emph=1)
foot(lb_ws, r + 3, "VAT (15%)", f"=ROUND(F{r+2}*{VAT},2)")
foot(lb_ws, r + 4, "LOT B — TOTAL (incl. VAT)", f"=F{r+2}+F{r+3}", emph=2)
lb_refs = {"sub": f"F{r}", "comm": f"F{r+1}", "ex": f"F{r+2}", "vat": f"F{r+3}", "inc": f"F{r+4}"}

lotb_notes = [
    "Note 1: Line B.1 – B.6 consolidates the complete turnkey pavilion scope of the RFP (fabrication & logistics, Block A & C build, Block B build, organizer approvals & clearances, main power supply and per-booth electrical hookups) as one all-inclusive package price — everything required to hand over a fully operational 162-SQM pavilion before the move-in deadline.",
    "Note 2: A dedicated Project Manager and on-site coordination team (including tea servers) are deployed from project start for exhibitor coordination, booth design/artwork liaison and troubleshooting.",
    "Note 3: All raw footage, edits and creative assets are delivered as exclusive intellectual property of PSEB.",
]
end = notes_block(lb_ws, r + 6, lotb_notes)
page(lb_ws, landscape=True, fit_h=1, titles=f"{HR}:{HR}", area=f"A1:H{end+1}")

# sanity: footer refs used on Summary must match reality
assert la_refs == LA, f"Lot A footer moved: {la_refs}"
assert lb_refs == LB, f"Lot B footer moved: {lb_refs}"

# =============================================================================
# 5. TEAM
# =============================================================================
tm = wb.create_sheet("Team")
set_widths(tm, [30, 66, 34])
NC = 3
band(tm, 1, NC, "RESOURCE PLAN & TALENT PROFILES",
     "Dedicated delivery team — single point of contact assigned exclusively to PSEB")

HR = 4
for j, htxt in enumerate(["Role", "Profile & Responsibility", "Deployment"], start=1):
    c = tm.cell(row=HR, column=j, value=htxt)
    c.font = f(9.5, True, WHITE)
    c.fill = fill(TEAL)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = box
tm.row_dimensions[HR].height = 20

team = [
    ("Project Director & Client Lead",
     "Overall engagement owner and commercial lead; direct escalation line for PSEB leadership; oversees both lots end-to-end.",
     "Full engagement, pre-event & on-site"),
    ("Dedicated Project Manager (single point of contact for PSEB)",
     "Assigned exclusively to PSEB for all pre-event and on-site liaison: exhibitor coordination, booth design & artwork approvals, schedule control, daily progress reporting and troubleshooting.",
     "Full-time from award to closeout"),
    ("Production Manager — Pavilion",
     "Leads fabrication, logistics and build of the 162-SQM pavilion; manages workshop, transport and RECC move-in/teardown per organizer schedule.",
     "Fabrication through teardown"),
    ("Site Supervisor & Safety Officer — RECC",
     "On-floor supervision of build crews; compliance with organizer stand audit, structural and HSE requirements.",
     "Move-in to teardown"),
    ("Certified Electrical Engineer",
     "Main power drops, panel deployment and per-booth hookups per Annex II/III load metrics; on-call through all live days.",
     "Build + 4 live days"),
    ("AV & Technical Lead — Forum",
     "SMD screens, mirror panels, sound and lighting systems at Crowne Plaza RDC; runs technical rehearsal and show calling.",
     "29 – 30 Aug"),
    ("Content Designer / Editor (2)",
     "Screen content per PSEB-approved event flow, daily highlight videos, social snippets and final documentaries.",
     "Pre-event + daily during events"),
    ("Photography Teams (2) & Videography Teams (2)",
     "Independent professional crews with permanent presence: ministerial bilaterals, speaker sessions, booth engagement and visitor traffic; Forum coverage on 30 Aug.",
     "30 Aug + 4 LEAP days"),
    ("Hospitality Team — Tea Servers & Lounge Attendants (4)",
     "Runs the tea/coffee station (approx. 500 cups/day), lounge dates & snacks service and daily water provisioning to all 20 booths.",
     "4 LEAP days"),
    ("Cleaning Crew (4)",
     "Continuous cleaning of pavilion, booths and lounge; dustbin emptying and waste removal.",
     "4 LEAP days"),
    ("On-site Event Coordinators (2)",
     "Guest registration, VVIP handling and hotel banquet liaison at the Business Forum; floor coordination at LEAP.",
     "30 Aug + 4 LEAP days"),
]
r = HR + 1
for i, (role, prof, dep) in enumerate(team):
    tm.cell(row=r, column=1, value=role).font = f(9.5, True, TEAL)
    tm.cell(row=r, column=2, value=prof).font = f(9.5)
    tm.cell(row=r, column=3, value=dep).font = f(9.5, False, MUTED)
    tm.cell(row=r, column=1).alignment = Alignment(vertical="center", wrap_text=True)
    tm.cell(row=r, column=2).alignment = Alignment(vertical="center", wrap_text=True)
    tm.cell(row=r, column=3).alignment = Alignment(vertical="center", wrap_text=True, horizontal="center")
    if i % 2 == 1:
        for col in range(1, NC + 1):
            tm.cell(row=r, column=col).fill = fill(TEAL_XLIGHT)
    for col in range(1, NC + 1):
        tm.cell(row=r, column=col).border = box
    tm.row_dimensions[r].height = max(desc_height(prof, cpl=75, line_h=12, extra=6, min_h=26),
                                      desc_height(role, cpl=34, line_h=12, extra=6, min_h=26))
    r += 1

rr = r + 1
tm.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=NC)
c = tm.cell(row=rr, column=1, value="CVs and detailed profiles of the assigned Project Manager and key personnel will be shared with PSEB upon award, prior to mobilization.")
c.font = f(8.5, False, MUTED, True)
c.alignment = Alignment(wrap_text=True)
page(tm, landscape=True, titles=f"{HR}:{HR}", area=f"A1:C{rr+1}")

# =============================================================================
# 6. INTERNAL (hidden) — costing & margin, not for client
# =============================================================================
iv = wb.create_sheet("Internal")
set_widths(iv, [9, 52, 15, 15, 13, 11, 42])
NC = 7
band(iv, 1, NC, "INTERNAL — COSTING & MARGIN (DO NOT PRINT / DO NOT SHARE)",
     "Quote cells pull live from the BOQ sheets; cost basis from supplier quotes as noted")

HR = 4
for j, htxt in enumerate(["Item", "Line", "Quote (SAR)", "Cost (SAR)", "Margin (SAR)", "Margin %", "Cost source"], start=1):
    c = iv.cell(row=HR, column=j, value=htxt)
    c.font = f(9.5, True, WHITE)
    c.fill = fill(TEAL)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = box

# Lot A sheet item rows: A.1=5, A.1.1=6, A.1.2=7, A.2=8, A.3=9, A.4=10, A.5=11,
# A.6=12, A.7=13, A.7.1=14, A.8=15, (A.9 included=16), A.10=17, A.11=18, A.12=19
lota_cost = [
    ("A.1", "Crowne Plaza package (hall + catering) — pass-through", 5, 325000, "Crowne Plaza RDC offer: 325,000 + 15% VAT, incl. hall space, dinner + 2 coffee breaks (priced as-is, no markup)"),
    ("A.1.1", "Event management & technical rehearsal", 6, 10000, "Innovation BOQ A.1 10,000"),
    ("A.1.2", "Stage 10m x 4m + carpet", 7, 7400, "Innovation BOQ A.8 5,600 + stage carpet 1,800"),
    ("A.2", "VVIP 50 sofas + tables + florals", 8, 12250, "Innovation BOQ A.2 subtotal 12,250"),
    ("A.3", "Media wall + red carpet + Q-poles", 9, 6500, "Innovation BOQ A.3 6,500"),
    ("A.4", "Registration desk + backdrop", 10, 8500, "Innovation BOQ A.4 8,500"),
    ("A.5", "Digital standees x4", 11, 3200, "Innovation BOQ A.5 4 x 800"),
    ("A.6", "Lighting + sound complete", 12, 15500, "Innovation BOQ A.6 10,000 + A.6.1 5,500"),
    ("A.7", "SMD main screen + server/washout", 13, 20500, "Innovation BOQ A.7 14,000 + A.7.1 6,500"),
    ("A.7.1", "SMD mirror panels x4", 14, 6300, "Est. ~350/sqm x 17.8 sqm + rigging (est.)"),
    ("A.8", "Branded podium + mic + stool", 15, 2800, "Market estimate (not in supplier BOQ)"),
    ("A.10", "On-site project coordinator", 17, 2500, "Internal staffing estimate"),
    ("A.11", "Media crew + full deliverables", 18, 6000, "Innovation BOQ A.10 4,500 + editing/deliverables est. 1,500"),
    ("A.12", "Ushers x10", 19, 6500, "Innovation BOQ A.11 10 x 650"),
]
# Lot B sheet item rows: B.1-B.6=5, B.7=6, B.8=7, B.9=8, B.10=9, B.11=10
lotb_cost = [
    ("B.1-B.6", "Turnkey pavilion (booth package)", 5, 130435, "Alpha Dimensions revised QUT-0126-26: 130,435 + VAT = 150,000 incl. VAT (orig. detailed quote 335,810 excl. VAT)"),
    ("B.7", "Tea & coffee station 4 days", 6, 6000, "Partially covered in Alpha revised (catering incl.); staff/consumables top-up est."),
    ("B.8", "Lounge dates & snacks 4 days", 7, 10000, "Catering supplier est. ~2,500/day"),
    ("B.9", "Water bottles 880", 8, 1320, "Est. 1.50/bottle wholesale"),
    ("B.10", "Cleaning & waste 4 days", 9, 4800, "Est. 1,200/day"),
    ("B.11", "2+2 media crews 4 days + editing", 10, 24000, "Partially covered in Alpha revised (coverage incl.); add'l crews/editing est. 6,000/day"),
]

def internal_row(r, code, label, quote_formula, cost, src):
    iv.cell(row=r, column=1, value=code).font = f(9.5, True, TEAL)
    iv.cell(row=r, column=2, value=label).font = f(9.5)
    iv.cell(row=r, column=3, value=quote_formula).number_format = "#,##0"
    iv.cell(row=r, column=4, value=cost).number_format = "#,##0"
    iv.cell(row=r, column=4).font = Font(name=FONT, size=9.5, color="0000FF")
    iv.cell(row=r, column=5, value=f"=C{r}-D{r}").number_format = "#,##0"
    iv.cell(row=r, column=6, value=f"=IF(D{r}=0,\"n/a\",C{r}/D{r}-1)").number_format = "0.0%"
    iv.cell(row=r, column=7, value=src).font = f(8.5, False, MUTED, True)
    iv.cell(row=r, column=7).alignment = Alignment(wrap_text=True, vertical="center")
    for col in range(1, NC + 1):
        iv.cell(row=r, column=col).border = box
    iv.row_dimensions[r].height = desc_height(src, cpl=44, line_h=11, extra=5, min_h=16)

r = HR + 1
iv_first = r
for code, label, lot_row, cost, src in lota_cost:
    internal_row(r, code, label, f"='Lot A - Forum'!F{lot_row}", cost, src)
    r += 1
for code, label, lot_row, cost, src in lotb_cost:
    internal_row(r, code, label, f"='Lot B - Pavilion'!F{lot_row}", cost, src)
    r += 1
iv_last = r - 1

iv.cell(row=r, column=2, value="TOTALS (excl. VAT, before agency commission)").font = f(10, True)
iv.cell(row=r, column=3, value=f"=SUM(C{iv_first}:C{iv_last})").number_format = "#,##0"
iv.cell(row=r, column=4, value=f"=SUM(D{iv_first}:D{iv_last})").number_format = "#,##0"
iv.cell(row=r, column=5, value=f"=C{r}-D{r}").number_format = "#,##0"
iv.cell(row=r, column=6, value=f"=IF(D{r}=0,\"n/a\",C{r}/D{r}-1)").number_format = "0.0%"
for col in range(1, NC + 1):
    iv.cell(row=r, column=col).border = box
    iv.cell(row=r, column=col).fill = fill(TEAL_LIGHT)
for col in (3, 4, 5, 6):
    iv.cell(row=r, column=col).font = f(10, True)
r += 1
iv.cell(row=r, column=2, value="Agency commission (12% on lot subtotals) — additional revenue on top").font = f(9.5)
iv.cell(row=r, column=3, value=f"='Lot A - Forum'!{LA['comm']}+'Lot B - Pavilion'!{LB['comm']}").number_format = "#,##0"
r += 2
iv.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NC)
c = iv.cell(row=r, column=1, value="Assumptions: Crowne Plaza RDC package 325,000 SAR + 15% VAT (hall + dinner + 2 coffee breaks, 300 pax) per offer relayed 18 Jul 2026, passed through at cost; supplier costs from Innovation Events BOQ (WhatsApp, 18 Jul 2026) and Alpha Dimensions QUT-0126-26 (20 Jul 2026); 'est.' items are internal estimates. Costs above are excl. VAT; input VAT on supplier invoices is recoverable.")
c.font = f(8.5, False, MUTED, True)
c.alignment = Alignment(wrap_text=True, vertical="top")
iv.row_dimensions[r].height = 40

page(iv, landscape=True, area=f"A1:G{r+1}")
iv.sheet_state = "hidden"

# =============================================================================
wb.properties.title = "PSEB LEAP 2026 — Commercial Proposal — Miradore Experiences"
wb.properties.creator = "Miradore Experiences"

out = "/home/user/miradore/leap2026/PSEB_LEAP2026_Commercial_Proposal_Miradore.xlsx"
wb.save(out)
print("saved", out)
